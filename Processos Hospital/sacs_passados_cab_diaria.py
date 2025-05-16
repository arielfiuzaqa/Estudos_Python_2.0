import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timedelta

def obter_caminho_planilha():
    """Obtém o caminho da planilha com a data de amanhã formatada."""
    tomorrow = datetime.now() + timedelta(days=0)
    date_tomorrow = tomorrow.strftime('%d-%m-%Y')
    caminho = f"C:\\Users\\t_ariel.fiuza\\Downloads\\PLANILHA DO CAB DIÁRIA - {date_tomorrow}.xlsx"
    return caminho

def carregar_planilha(caminho):
    """Carrega a planilha do caminho especificado."""
    try:
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active
        return wb, ws
    except FileNotFoundError:
        print(f"Erro: Arquivo não encontrado em {caminho}.")
        return None, None

def obter_sacs():
    """Obtém os SAC's do usuário e retorna uma lista."""
    sacs_input = input("Digite os SAC's separados por vírgula: ")
    sacs = [sac.strip() for sac in sacs_input.split(',')]
    return sacs

def definir_estilo_celula():
    """Define o estilo da célula 'CAB Passado'."""
    fill_yellow = PatternFill(start_color="3F00FF", end_color="3F00FF", fill_type="solid") # Mantem o azul escuro
    font_white_bold = Font(color="FFFFFF", bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    return fill_yellow, font_white_bold, alignment_center

def marcar_semanal_passada(ws, sacs, estilo):
    """Marca a célula 'CAB Passado' na coluna A para os SAC's encontrados na coluna B."""
    fill_yellow, font_white_bold, alignment_center = estilo
    
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):  # Coluna B, a partir da linha 2
        cell = row[0]
        if str(cell.value) in sacs:
            gap_cell = ws.cell(row=cell.row, column=1)
            gap_cell.value = "CAB Passado"
            gap_cell.fill = fill_yellow
            gap_cell.font = font_white_bold
            gap_cell.alignment = alignment_center

def salvar_planilha(wb, caminho):
    """Salva a planilha com as alterações no caminho especificado."""
    wb.save(caminho)
    print("Planilha salva com sucesso.")

def main():
    caminho = obter_caminho_planilha()
    wb, ws = carregar_planilha(caminho)
    if wb and ws:
        sacs = obter_sacs()
        estilo = definir_estilo_celula()
        marcar_semanal_passada(ws, sacs, estilo)
        salvar_planilha(wb, caminho)

if __name__ == "__main__":
    main()
