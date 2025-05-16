import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime, timedelta

# Define o caminho da pasta de Downloads
pasta_downloads = "C:\\Users\\t_ariel.fiuza\\Downloads\\"

# Definindo o nome da aba como uma variável
nome_aba = 'SAC\'s'  # Nome atual da aba que você deseja modificar
nome_aba_semanal = 'SAC\'s CAB Diária'  # Usado para a aba temporária
nome_aba_baixa = 'PARA DAR BAIXA'  # Usado para a aba temporária da BAIXA

# Caminho da planilha original
caminho_planilha_original = pasta_downloads + "Planilha geral de Publicação Online!.xlsx"

# Calculo de D+1 para a planilha de publicação
tomorrow = datetime.now() + timedelta(days=0)
date_tomorrow = tomorrow.strftime('%d-%m-%Y')

# Nome do arquivo com data atual, salvando na pasta de Downloads
# nome_nova_planilha = f"{pasta_downloads}PLANILHA DE PUBLICAÇÃO SEMANAL - {datetime.now().strftime('%d-%m-%Y')}.xlsx"
nome_nova_planilha = f"{pasta_downloads}PLANILHA DO CAB DIÁRIA - {date_tomorrow}.xlsx"

# Abre a planilha específica (SAC's) e filtra os dados
df = pd.read_excel(caminho_planilha_original, sheet_name=nome_aba, dtype={"Versão": str})
df_filtrado = df[df['STATUS'] == 'Publicação Semanal'].iloc[:, :13]
df_filtrado['STATUS'] = df_filtrado['STATUS'].replace("Publicação Semanal", "CAB Diária")
df_filtrado_baixa = df[(df['STATUS'] == 'Enviado Emergencial') | (df['STATUS'] == 'Já estava em produção') | (df['STATUS'] == 'Enviado Padrão')].iloc[:, :13]

# Ajusta o formato dos números na coluna 'Versão'
df_filtrado['Versão'] = df_filtrado['Versão'].apply(lambda x: str(x).replace(',', '.') if pd.notnull(x) else x)
df_filtrado_baixa['Versão'] = df_filtrado_baixa['Versão'].apply(lambda x: str(x).replace(',', '.') if pd.notnull(x) else x)

# **Adicione a ordenação aqui**
df_filtrado.sort_values(by=['TORRE', 'PO Aprovador', 'Coordenador', 'Analista', 'SAC', 'Objeto', 'Tipo de Objeto', 'Versão'], inplace=True)
df_filtrado_baixa.sort_values(by=['TORRE', 'PO Aprovador', 'Coordenador', 'Analista', 'SAC', 'Objeto', 'Tipo de Objeto', 'Versão'], inplace=True)

# Salva o DataFrame filtrado usando pandas na aba semanal + Status 'CAB DIÁRIA'
df_filtrado.to_excel(nome_nova_planilha, index=False, sheet_name=nome_aba_semanal)

# Carrega o workbook existente com openpyxl para adicionar o segundo DataFrame
book = load_workbook(nome_nova_planilha)

# Agora, usar o ExcelWriter com a engine do openpyxl e passando o livro carregado como argumento
with pd.ExcelWriter(nome_nova_planilha, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    book = writer.book
    df_filtrado_baixa.to_excel(writer, index=False, sheet_name=nome_aba_baixa)
# Note que 'mode='a'' é usado para abrir o arquivo existente para adição

# Carrega a nova planilha criada com openpyxl para editar estilos
wb = load_workbook(nome_nova_planilha)
ws = wb[nome_aba_semanal]  # Para aplicar os estilos na aba 'SAC's Semanal'
ws_baixa = wb[nome_aba_baixa]  # Para aplicar os estilos na aba 'PARA DAR BAIXA'

# (Código para definir e aplicar os estilos segue exatamente como você forneceu)
# Define estilos
cor_fundo_cabecalho = "030303"   # Preto
cor_fonte_cabecalho = "FFFFFF"  # Branco
cor_fundo_publicacao = "ED7D31"  # Laranja
cor_fundo_emergencial = "009900" # Verde Escuro
cor_fonte_publicacao = "FFFFFF"  # Branco
borda_preta = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                     top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
centralizado = Alignment(horizontal='center', vertical='center')

# Aplica os estilos ao cabeçalho e às células conforme especificado
for col in range(1, 14):  # De A (1) até M (13)
    celula = ws.cell(row=1, column=col)
    celula.fill = PatternFill(start_color=cor_fundo_cabecalho, end_color=cor_fundo_cabecalho, fill_type="solid")
    celula.font = Font(color=cor_fonte_cabecalho, bold=True)
    celula.border = borda_preta
    celula.alignment = centralizado

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=13):
    for cell in row:
        if cell.column_letter in ['A', 'B', 'C', 'D', 'F', 'K', 'L']:
            cell.alignment = centralizado
        if cell.column_letter == 'A' and cell.value == "CAB Diária":
            cell.fill = PatternFill(start_color=cor_fundo_publicacao, end_color=cor_fundo_publicacao, fill_type="solid")
            cell.font = Font(color=cor_fonte_publicacao, bold=True)
        cell.border = borda_preta

# Continuação do código para aplicar estilos na aba 'PARA DAR BAIXA'
# Aplica os estilos ao cabeçalho na aba 'PARA DAR BAIXA'
for col in range(1, ws_baixa.max_column + 1):
    celula = ws_baixa.cell(row=1, column=col)
    celula.fill = PatternFill(start_color=cor_fundo_cabecalho, end_color=cor_fundo_cabecalho, fill_type="solid")
    celula.font = Font(color=cor_fonte_cabecalho, bold=True)
    celula.border = borda_preta
    celula.alignment = centralizado

# Aplica os estilos às células de dados na aba 'PARA DAR BAIXA'
for row in ws_baixa.iter_rows(min_row=2, max_row=ws_baixa.max_row, min_col=1, max_col=ws_baixa.max_column):
    for cell in row:
        cell.border = borda_preta
        cell.alignment = centralizado
        # Verifica se a célula está na coluna 'A' e se o valor corresponde à condição para aplicar o estilo de publicação
        if cell.column_letter == 'A' and (cell.value == "Enviado Emergencial" or cell.value == "Já estava em produção" or cell.value == "Enviado Padrão"):
            cell.fill = PatternFill(start_color=cor_fundo_emergencial, end_color=cor_fundo_emergencial, fill_type="solid")
            cell.font = Font(color=cor_fonte_publicacao, bold=True)

# Ajuste para alinhamento à esquerda nas colunas E, G, H, I e J
for row in ws_baixa.iter_rows(min_row=2, max_row=ws_baixa.max_row):
    for cell in row:
        if cell.column in [5, 7, 8, 9, 10]:  # Correspondendo às colunas E (5), G (7), H (8), I (9) e J (10)
            cell.alignment = Alignment(horizontal='left')

# Função para ajustar as cores das versões conforme especificado
def ajustar_cores_versoes():
    versoes = {}
    for row in ws.iter_rows(min_row=2):
        id_item = (row[3].value, row[4].value)  # Colunas D e E como identificador - Modificar aqui ***
        homolog = row[2].value  # Coluna C Homologação
        versao = row[5].value  # Coluna F Versão
        if id_item not in versoes:
            versoes[id_item] = []
        versoes[id_item].append((homolog, versao, row[5]))  # Adiciona homologação, versão e a célula da versão

    azul_escuro = "002060"
    azul_claro = "00B0F0"
    branco = "FFFFFF"

    # Itera pelos itens para encontrar a maior versão e homologação
    for id_item, versoes_item in versoes.items():
        if len(versoes_item) > 1:
            # Ordena primeiro por homologação e depois por versão
            versoes_item.sort(reverse=True, key=lambda x: (x[0], x[1])) 
            maior_combinacao = versoes_item[0][2]  # A célula com a maior combinação de homologação e versão

            # Aplica cores
            for homolog, versao, celula in versoes_item:
                if celula == maior_combinacao:
                    celula.fill = PatternFill(start_color=azul_escuro, end_color=azul_escuro, fill_type="solid")
                    celula.font = Font(color=branco, bold=True)
                else:
                    celula.fill = PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type="solid")
                    celula.font = Font(color=azul_escuro, bold=True)

ajustar_cores_versoes()

# Salva as alterações na planilha
wb.save(nome_nova_planilha)
print(f"Planilha '{nome_nova_planilha}' atualizada e salva com sucesso.")