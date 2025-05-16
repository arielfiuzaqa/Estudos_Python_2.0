import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Função para remover extensão de um objeto
def remover_extensao(nome_objeto):
    if isinstance(nome_objeto, str):
        return nome_objeto.split('.')[0].lower()  # Converte para minúscula
    return nome_objeto

# Caminho para a pasta Downloads
pasta_downloads = "C:\\Users\\t_ariel.fiuza\\Downloads\\"

# Nome dos arquivos com base na data de amanhã
tomorrow = datetime.now() + timedelta(days=0)
date_tomorrow = tomorrow.strftime('%d-%m-%Y')

# Caminhos dos arquivos
arquivo_planilha_semanal = pasta_downloads + f"PLANILHA DO CAB DIÁRIA - {date_tomorrow}.xlsx"
arquivo_planilha_war = pasta_downloads + "Objetos War.xlsx"

# Carregar as planilhas
planilha_semanal = pd.read_excel(arquivo_planilha_semanal)
planilha_war = pd.read_excel(arquivo_planilha_war)

# Remover extensões e converter para minúsculas para a comparação
coluna_semanal = planilha_semanal['Objeto'].apply(remover_extensao)
coluna_war = planilha_war['WAR'].apply(remover_extensao) # Add JAR e ZIP para verificar essas extensões

# Abrir o arquivo com openpyxl para edição
workbook = load_workbook(arquivo_planilha_semanal)
sheet = workbook.active

# Definir o preenchimento em vermelho para as células
fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Contagem de comparações feitas e itens encontrados
comparacoes = 0
itens_encontrados = []

# Realizar a comparação e pintar as células iguais de vermelho
for index, objeto in enumerate(coluna_semanal):
    if objeto in coluna_war.values:
        # Pintar a célula de vermelho na coluna E (index + 2 por causa do cabeçalho)
        cell = sheet[f'E{index + 2}']
        cell.fill = fill_red
        comparacoes += 1
        itens_encontrados.append(objeto)  # Adicionar o objeto à lista de encontrados

# Salvar as alterações na planilha
workbook.save(arquivo_planilha_semanal)

# Exibir o resultado no terminal
print(f"Comparação finalizada! {comparacoes} itens iguais encontrados e pintados de vermelho na planilha 'TAB PLANILHA DE PUBLICAÇÃO SEMANAL - {date_tomorrow}.xlsx'.")
if itens_encontrados:
    print("Itens encontrados:")
    for item in set(itens_encontrados):  # Use set para listar cada item apenas uma vez
        print(item)
else:
    print("Nenhum item encontrado.")
